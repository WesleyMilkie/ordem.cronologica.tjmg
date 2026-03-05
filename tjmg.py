from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException
from webdriver_manager.chrome import ChromeDriverManager
import os
import time
from tabulate import tabulate
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
import psycopg2
from psycopg2 import sql
from psycopg2.extras import execute_batch
from dotenv import load_dotenv
import re


def encontrar_primeiro_elemento(driver, seletores, timeout=5):
    """
    Tenta localizar um elemento por uma lista de seletores (By, valor).
    Retorna o primeiro elemento encontrado ou None.
    """
    for by, valor in seletores:
        try:
            return WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((by, valor))
            )
        except Exception:
            continue
    return None


def extrair_total_precatorios(driver):
    """
    Extrai o total de precatórios do elemento de paginação.
    Tenta múltiplos locais onde o elemento pode estar.
    Usa regex para extrair o número formatado com separadores de milhar.
    Retorna o número extraído ou None se não encontrado.
    """
    try:
        texto = None
        elemento = None

        seletores_footer = [
            (
                By.CSS_SELECTOR,
                "#resultado > div.ui-datatable-footer.ui-widget-header.ui-corner-bottom",
            ),
            (
                By.XPATH,
                '//*[@id="resultado"]//div[contains(@class,"ui-datatable-footer")]',
            ),
            (By.XPATH, '//*[@id="resultado"]/div[5]'),
            (By.XPATH, '//*[@id="resultado"]/div[3]'),
        ]

        elemento = encontrar_primeiro_elemento(driver, seletores_footer, timeout=3)
        if elemento:
            texto = (elemento.text or "").strip()
            if texto:
                print("  ✓ Elemento de paginação encontrado")

        if not texto:
            print(f"  ⚠ Elemento de paginação não encontrado")
            return None

        print(f"  📊 Texto do paginador: {texto}")

        # Padrão 1: "de NÚMERO registros" ou "de NÚMERO precatórios"
        # Aceita números com separadores de milhar (pontos): "de 14.293 precatórios"
        match = re.search(
            r"de\s+([\d\.]+)\s+(registros?|precatórios?)", texto, re.IGNORECASE
        )
        if match:
            # Remover pontos (separadores de milhar) e converter para int
            numero_str = match.group(1).replace(".", "")
            total = int(numero_str)
            print(f"  ✓ Total extraído (padrão 'de X precatórios'): {total}")
            return total

        # Padrão 2: "Total de NÚMERO precatórios."
        match = re.search(r"total\s+de\s+([\d\.]+)", texto, re.IGNORECASE)
        if match:
            numero_str = match.group(1).replace(".", "")
            total = int(numero_str)
            print(f"  ✓ Total extraído (padrão 'Total de X'): {total}")
            return total

        # Padrão alternativo: pegar o último número (que geralmente é o total)
        # Aceita números com pontos como separadores de milhar
        numeros = re.findall(r"[\d\.]+", texto)
        if numeros:
            # Pegar o último número (geralmente é o total de registros/precatórios)
            numero_str = numeros[-1].replace(".", "")
            total = int(numero_str)
            print(f"  ✓ Número extraído (último encontrado): {total}")
            return total

        print(f"  ⚠ Não foi possível extrair número do texto: {texto}")
        return None

    except Exception as e:
        print(f"  ⚠ Erro ao extrair total de precatórios: {e}")
        return None


def salvar_validacao(
    entidade_num,
    ente_devedor_principal,
    ente_devedor,
    total_precatorios,
    worker_id=None,
):
    """
    Salva dados de validação na tabela tjmg_validacao.
    """
    try:
        conn = obter_conexao_postgres()
        conn.autocommit = True

        with conn.cursor() as cur:
            cur.execute(
                sql.SQL(
                    """
                    INSERT INTO ordens_cronologicas.tjmg_validacao 
                    (entidade_num, ente_devedor_principal, ente_devedor, total_precatorios_site, worker_id)
                    VALUES (%s, %s, %s, %s, %s)
                    ON CONFLICT (entidade_num) 
                    DO UPDATE SET 
                        ente_devedor_principal = EXCLUDED.ente_devedor_principal,
                        ente_devedor = EXCLUDED.ente_devedor,
                        total_precatorios_site = EXCLUDED.total_precatorios_site,
                        data_extracao = CURRENT_TIMESTAMP,
                        worker_id = EXCLUDED.worker_id
                    """
                ),
                (
                    entidade_num,
                    ente_devedor_principal,
                    ente_devedor,
                    total_precatorios,
                    worker_id,
                ),
            )

        conn.close()
        print(
            f"  ✓ Validação salva: Entidade #{entidade_num} - {total_precatorios} precatórios"
        )

    except Exception as e:
        print(f"  ⚠ Erro ao salvar validação: {e}")


def extrair_informacoes_modal(driver):
    """
    Extrai as informações do modal aberto usando XPaths específicos.
    Retorna um dicionário com os dados tabulares.
    """
    informacoes = {
        "Ente Devedor Principal": "",
        "Ente Devedor": "",
        "Regime de Pagamento": "",
        "Lei Pequeno Valor": "",
        "ordem_cronologica_nao_disponivel": False,  # FLAG para controle posterior
    }

    try:
        # Aguardar um pouco para garantir que o modal carregou
        time.sleep(1)

        # Estratégia 1: Usar XPath direto para "Ente Principal"
        try:
            elemento_ente_principal = driver.find_element(
                By.XPATH, '//*[@id="entePrincipal"]'
            )
            valor = (
                elemento_ente_principal.get_attribute("value")
                or elemento_ente_principal.text
            )
            informacoes["Ente Devedor Principal"] = valor.strip() if valor else ""
            print(
                f"✓ Ente Devedor Principal extraído: {informacoes['Ente Devedor Principal']}"
            )
        except Exception as e:
            print(f"⚠ Não foi possível extrair Ente Principal: {e}")

        # Estratégia 1b: Usar XPath direto para "Ente Devedor"
        try:
            elemento_ente_devedor = driver.find_element(
                By.XPATH, '//*[@id="enteDevedor"]'
            )
            valor = (
                elemento_ente_devedor.get_attribute("value")
                or elemento_ente_devedor.text
            )
            informacoes["Ente Devedor"] = valor.strip() if valor else ""
            print(f"✓ Ente Devedor extraído: {informacoes['Ente Devedor']}")
        except Exception as e:
            print(f"⚠ Não foi possível extrair Ente Devedor: {e}")

        # Estratégia 1c: Usar XPath direto para "Regime de Pagamento"
        try:
            elemento_regime = driver.find_element(
                By.XPATH, '//*[@id="regimePagamento"]'
            )
            valor = elemento_regime.get_attribute("value") or elemento_regime.text
            informacoes["Regime de Pagamento"] = valor.strip() if valor else ""
            print(
                f"✓ Regime de Pagamento extraído: {informacoes['Regime de Pagamento']}"
            )
        except Exception as e:
            print(f"⚠ Não foi possível extrair Regime de Pagamento: {e}")

        # Estratégia 1d: Usar XPath direto para "Lei Pequeno Valor"
        try:
            elemento_lei_pequeno = driver.find_element(
                By.XPATH, '//*[@id="leiPequenoValor"]'
            )
            valor = (
                elemento_lei_pequeno.get_attribute("value") or elemento_lei_pequeno.text
            )
            informacoes["Lei Pequeno Valor"] = valor.strip() if valor else ""
            print(f"✓ Lei Pequeno Valor extraído: {informacoes['Lei Pequeno Valor']}")
        except Exception as e:
            print(f"⚠ Não foi possível extrair Lei Pequeno Valor: {e}")

        # Estratégia 2: Procurar por labels com texto específico (apenas para campos não cobertos)
        labels = driver.find_elements(By.TAG_NAME, "label")

        for label in labels:
            texto_label = label.text.strip().lower()

            # Evitar duplicação - pula se já foi extraído via XPath
            if any(
                x in texto_label
                for x in [
                    "ente devedor principal",
                    "ente devedor",
                    "regime de pagamento",
                    "lei pequeno valor",
                    "pequeno valor",
                ]
            ):
                continue

            # Encontrar o campo de entrada/texto associado ao label
            try:
                # Procurar dentro do mesmo container
                parent = label.find_element(By.XPATH, "..")

                # Tenta encontrar input primeiro
                try:
                    valor_element = parent.find_element(By.TAG_NAME, "input")
                    valor = valor_element.get_attribute("value") or valor_element.text
                except:
                    # Se não encontrar input, tenta span
                    valor_element = parent.find_element(By.TAG_NAME, "span")
                    valor = valor_element.text

            except Exception as e:
                pass

        # Estratégia 3: Procurar por parsing de texto como fallback final
        try:
            corpo_modal = driver.find_element(By.TAG_NAME, "body")
            texto_completo = corpo_modal.text
            linhas = texto_completo.split("\n")

            for i, linha in enumerate(linhas):
                linha_lower = linha.lower().strip()

                # Apenas como fallback se não foi extraído via XPath
                if (
                    "ente devedor:" in linha_lower
                    and i + 1 < len(linhas)
                    and not informacoes["Ente Devedor"]
                ):
                    informacoes["Ente Devedor"] = linhas[i + 1].strip()
        except Exception as e:
            pass

        # Verificar se a mensagem "Ordem Cronológica de Pagamento não disponível" está presente
        try:
            elemento_mensagem = encontrar_primeiro_elemento(
                driver,
                [
                    (By.XPATH, '//*[@id="j_idt23"]/div[1]/ul/li/span'),
                    (
                        By.XPATH,
                        '//*[contains(normalize-space(.),"Ordem Cronológica de Pagamento não disponível nesta consulta!")]',
                    ),
                ],
                timeout=2,
            )
            if not elemento_mensagem:
                raise Exception("mensagem não encontrada")
            texto_mensagem = elemento_mensagem.text.strip()

            if (
                "Ordem Cronológica de Pagamento não disponível nesta consulta!"
                in texto_mensagem
            ):
                informacoes["ordem_cronologica_nao_disponivel"] = True
                print(f"⚠ ⚠ AVISO: {texto_mensagem}")
                print(f"🚩 FLAG 'ordem_cronologica_nao_disponivel' definida como TRUE")
            else:
                informacoes["ordem_cronologica_nao_disponivel"] = False
        except:
            # XPath não encontrado, significa que a mensagem não está presente
            informacoes["ordem_cronologica_nao_disponivel"] = False

    except Exception as e:
        print(f"Erro ao extrair informações do modal: {e}")

    return informacoes


def exibir_informacoes_tabular(informacoes, numero_entidade):
    """
    Exibe as informações extraídas em formato tabular no terminal.
    """
    print(f"\n{'='*70}")
    print(f"📊 INFORMAÇÕES DA ENTIDADE #{numero_entidade}")
    print(f"{'='*70}")

    dados_tabela = [
        ["Ente Devedor Principal", informacoes.get("Ente Devedor Principal", "N/A")],
        ["Ente Devedor", informacoes.get("Ente Devedor", "N/A")],
        ["Regime de Pagamento", informacoes.get("Regime de Pagamento", "N/A")],
        ["Lei Pequeno Valor", informacoes.get("Lei Pequeno Valor", "N/A")],
        [
            "Status Ordem Cronológica",
            (
                "❌ NÃO DISPONÍVEL"
                if informacoes.get("ordem_cronologica_nao_disponivel")
                else "✅ DISPONÍVEL"
            ),
        ],
    ]

    print(tabulate(dados_tabela, headers=["Campo", "Valor"], tablefmt="grid"))
    print(f"{'='*70}\n")


def extrair_cabecalho_tabela(driver):
    """Extrai os nomes das colunas do cabeçalho da tabela resultado."""
    headers = []
    try:
        thead = driver.find_element(By.XPATH, '//*[@id="resultado_head"]')
        ths = thead.find_elements(By.TAG_NAME, "th")
        headers = [th.text.strip() for th in ths if th.text.strip()]
        print(f"  ✓ Cabeçalhos extraídos: {headers}")
    except Exception as e:
        print(f"  ⚠ Erro ao extrair cabeçalhos: {e}")
    return headers


def extrair_subtabela(driver, row_index):
    """Extrai dados da subtabela expandida para uma linha específica."""
    dados_sub = {"headers": [], "linhas": []}

    xpaths_candidatos = [
        f'//*[@id="resultado:{row_index}:j_idt45_content"]',
        f'//*[@id="resultado:{row_index}:j_idt46:display"]',
    ]

    subtabela_container = None
    for xp in xpaths_candidatos:
        try:
            subtabela_container = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, xp))
            )
            break
        except Exception:
            continue

    if not subtabela_container:
        print(f"    ⚠ Subtabela não encontrada para linha {row_index}")
        return dados_sub

    try:
        # Tentar encontrar uma tabela dentro do container
        try:
            tabela = subtabela_container.find_element(By.TAG_NAME, "table")
        except Exception:
            tabela = subtabela_container

        # Extrair cabeçalhos da subtabela
        try:
            sub_headers = tabela.find_elements(By.TAG_NAME, "th")
            dados_sub["headers"] = [
                h.text.strip() for h in sub_headers if h.text.strip()
            ]
        except Exception:
            pass

        # Extrair linhas da subtabela
        try:
            linhas_sub = tabela.find_elements(By.TAG_NAME, "tr")
        except Exception:
            linhas_sub = []

        for linha in linhas_sub:
            colunas = linha.find_elements(By.TAG_NAME, "td")
            if colunas:
                row_data = [col.text.strip() for col in colunas]
                dados_sub["linhas"].append(row_data)

        print(f"    ✓ Subtabela: {len(dados_sub['linhas'])} linhas extraídas")
        if dados_sub["headers"]:
            print(f"    ✓ Subtabela headers: {dados_sub['headers']}")
    except Exception as e:
        print(f"    ⚠ Erro ao extrair subtabela da linha {row_index}: {e}")

    return dados_sub


def parsear_andamento(texto):
    """Faz parsing do texto de andamento para separar em campos individuais."""
    campos = {}

    if not texto:
        return campos

    # Mapear labels para nomes de colunas limpos
    mapeamento = {
        "Credor Principal:": "Credor Principal",
        "Número e Natureza do Precatório:": "Número e Natureza do Precatório",
        "Ano de Vencimento:": "Ano de Vencimento",
        "Processo Eproc 2ª Instância nº:": "Processo Eproc 2ª Instância",
        "Situação:": "Situação",
        "Valor de formação do Precatório (Valor de Face):": "Valor de Face",
        "Data da última atualização do Valor de Face (Data de Liquidação):": "Data de Liquidação",
        "Protocolo (Data/Hora):": "Protocolo Data/Hora",
        "Protocolo (Número/Ano):": "Protocolo Número/Ano",
        "Processo de Execução nº:": "Processo de Execução",
        "Processo SEI nº:": "Processo SEI",
        "Origem:": "Origem",
        "Ação:": "Ação",
        "Ordem Cronológica": "Ordem Cronológica",
        "Ordem Aberto/Suspenso?": "Ordem Aberto/Suspenso",
    }

    # Separar por linhas
    linhas = texto.split("\n")

    for linha in linhas:
        linha = linha.strip()
        if not linha:
            continue

        # Tentar encontrar o padrão "Label: Valor"
        encontrou = False
        for label_original, nome_coluna in mapeamento.items():
            if linha.startswith(label_original):
                valor = linha.replace(label_original, "").strip()
                campos[nome_coluna] = valor
                encontrou = True
                break

        # Se não encontrou com ":", tentar match exato (para campos sem valor como "Ordem Cronológica")
        if not encontrou:
            for label_original, nome_coluna in mapeamento.items():
                # Remove ":" do label para comparar
                label_sem_dois_pontos = label_original.rstrip(":")
                if linha == label_sem_dois_pontos or linha == label_original:
                    campos[nome_coluna] = "Sim"  # Marca presença do campo
                    break

    return campos


def parsear_subtabela_linhas(linhas_sub):
    """Converte linhas da subtabela em dicionário chave/valor quando possível."""
    dados = {}
    for linha in linhas_sub:
        if not linha:
            continue

        # Caso comum: ["Campo", "Valor"]
        if len(linha) >= 2:
            chave = (linha[0] or "").strip()
            valor = (linha[1] or "").strip()
            if chave:
                dados[chave] = valor
            continue

        # Fallback: tentar dividir "Campo = Valor" em uma única célula
        if len(linha) == 1:
            texto = (linha[0] or "").strip()
            if "=" in texto:
                partes = texto.split("=", 1)
                chave = partes[0].strip()
                valor = partes[1].strip()
                if chave:
                    dados[chave] = valor

    return dados


def extrair_andamento_e_beneficiarios(driver, wait, row_index):
    """Clica em idAndamento e extrai detalhes + beneficiários."""
    detalhes = {
        "andamento": "",
        "beneficiarios": "",
        "andamento_completo": "",
        "andamento_campos": {},
    }

    # Primeiro, verificar se o botão de andamento existe
    linha_local = (row_index % 15) + 1
    try:
        btn_andamento = encontrar_primeiro_elemento(
            driver,
            [
                (By.XPATH, f'//*[@id="resultado:{row_index}:idAndamento"]'),
                (
                    By.XPATH,
                    f'//*[@id="resultado_data"]/tr[{linha_local}]//*[contains(@id,":idAndamento")]',
                ),
                (
                    By.XPATH,
                    f'//*[@id="resultado_data"]/tr[{linha_local}]//a[contains(translate(normalize-space(.), "ANDAMENTO", "andamento"), "andamento")]',
                ),
            ],
            timeout=3,
        )
        if not btn_andamento:
            raise Exception("botão idAndamento não encontrado")
    except:
        print(
            f'    ❌❌❌ ATENÇÃO: LINHA {row_index} NÃO POSSUI BOTÃO idAndamento - XPATH: //*[@id="resultado:{row_index}:idAndamento"] ❌❌❌'
        )
        return detalhes

    try:
        # Clicar no botão de andamento
        driver.execute_script("arguments[0].scrollIntoView(true);", btn_andamento)
        time.sleep(0.5)
        driver.execute_script("arguments[0].click();", btn_andamento)
        print(f"    🔗 Clicou em idAndamento da linha {row_index}")
        time.sleep(4)

        # Extrair APENAS o conteúdo do xpath j_idt104 (SEM a Observação Importante)
        try:
            # Extrair apenas o elemento j_idt104 (painel principal de detalhes)
            try:
                painel_detalhes = encontrar_primeiro_elemento(
                    driver,
                    [
                        (By.XPATH, '//*[@id="frm_detalhe:j_idt91:j_idt104"]'),
                        (
                            By.XPATH,
                            '//*[@id="frm_detalhe"]//*[contains(@id,":j_idt104")]',
                        ),
                        (
                            By.XPATH,
                            '//*[@id="frm_detalhe"]//*[contains(normalize-space(.),"Credor Principal:")]',
                        ),
                    ],
                    timeout=10,
                )
                if not painel_detalhes:
                    raise Exception("painel de detalhes não encontrado")
                texto_detalhes = painel_detalhes.text.strip()

                # Remover tudo após "=== Observação Importante ===" caso esteja no texto
                if "=== Observação Importante ===" in texto_detalhes:
                    texto_detalhes = texto_detalhes.split(
                        "=== Observação Importante ==="
                    )[0].strip()

                # Fazer parsing do texto para extrair campos individuais
                detalhes["andamento_completo"] = texto_detalhes
                detalhes["andamento_campos"] = parsear_andamento(texto_detalhes)
                detalhes["andamento"] = texto_detalhes  # Manter compatibilidade
                print(
                    f"    ✓ Andamento extraído e parseado (j_idt104): {len(texto_detalhes)} chars, {len(detalhes['andamento_campos'])} campos"
                )
            except Exception as e:
                print(f"    ⚠ Erro ao extrair j_idt104: {e}")
                detalhes["andamento"] = ""
                detalhes["andamento_completo"] = ""
                detalhes["andamento_campos"] = {}
        except Exception as e:
            print(f"    ⚠ Erro ao extrair andamento: {e}")
            detalhes["andamento"] = ""
            detalhes["andamento_completo"] = ""
            detalhes["andamento_campos"] = {}

        # Extrair beneficiários
        try:
            beneficiarios_el = encontrar_primeiro_elemento(
                driver,
                [
                    (
                        By.XPATH,
                        '//*[@id="frm_detalhe:j_idt91:crontrol_lblBeneficiario"]',
                    ),
                    (
                        By.XPATH,
                        '//*[@id="frm_detalhe"]//*[contains(@id,"Beneficiario")]',
                    ),
                    (
                        By.XPATH,
                        '//*[@id="frm_detalhe"]//*[contains(normalize-space(.),"Beneficiários")]',
                    ),
                ],
                timeout=3,
            )
            if not beneficiarios_el:
                raise Exception("bloco de beneficiários não encontrado")
            detalhes["beneficiarios"] = beneficiarios_el.text.strip()
            print(f"    ✓ Beneficiários: {detalhes['beneficiarios'][:100]}")
        except Exception as e:
            print(f"    ⚠ Erro ao extrair beneficiários: {e}")

        # Fechar o modal de detalhes para continuar nas outras linhas
        try:
            btn_fechar = encontrar_primeiro_elemento(
                driver,
                [
                    (By.XPATH, '//*[@id="frm_detalhe:j_idt91:j_idt100"]/span'),
                    (
                        By.CSS_SELECTOR,
                        "#frm_detalhe a.ui-dialog-titlebar-close, #frm_detalhe span.ui-icon-closethick",
                    ),
                    (
                        By.XPATH,
                        '//*[@id="frm_detalhe"]//*[contains(@class,"ui-dialog-titlebar-close") or contains(@class,"ui-icon-closethick")]',
                    ),
                ],
                timeout=3,
            )
            if not btn_fechar:
                raise Exception("botão fechar modal não encontrado")
            driver.execute_script("arguments[0].click();", btn_fechar)
            print("    ✓ Modal de detalhes fechado")
            time.sleep(1.5)
        except Exception as e:
            print(f"    ⚠ Não foi possível fechar o modal: {e}")

    except Exception as e:
        print(f"    ⚠ Erro ao acessar andamento da linha {row_index}: {e}")

    return detalhes


def ir_proxima_pagina(driver):
    """Tenta navegar para a próxima página. Retorna False se não houver mais páginas."""
    try:
        paginador = None
        for xp in [
            '//*[@id="resultado_paginator_top"]',
            '//*[@id="resultado_paginator_bottom"]',
        ]:
            try:
                paginador = driver.find_element(By.XPATH, xp)
                break
            except Exception:
                continue

        if not paginador:
            print("  📄 Paginador não encontrado; assumindo página única")
            return False
        # PrimeFaces usa .ui-paginator-next como botão de próxima página
        btn_next = paginador.find_element(By.CSS_SELECTOR, ".ui-paginator-next")

        # Verificar se o botão está desabilitado
        classes = btn_next.get_attribute("class") or ""
        if "ui-state-disabled" in classes:
            print("  📄 Última página alcançada")
            return False

        btn_next.click()
        time.sleep(3)
        return True
    except Exception as e:
        print(f"  📄 Sem mais páginas ou erro na paginação: {e}")
        return False


def extrair_tabela_resultado(
    driver,
    wait,
    info_entidade,
    iniciar_de_pagina=1,
    iniciar_de_linha=0,
    callback_checkpoint=None,
):
    """
    Extrai todos os dados da tabela resultado, incluindo subtabelas,
    andamentos e beneficiários, com suporte a paginação.

    Args:
        iniciar_de_pagina: Página a partir da qual começar a extração (default: 1)
        iniciar_de_linha: Linha (0-based) para iniciar na página inicial (default: 0)
        callback_checkpoint: Função para salvar checkpoint após cada página (pagina_num)
    """
    resultado = {
        "info_entidade": info_entidade,
        "headers": [],
        "registros": [],
        "_pagina_atual": iniciar_de_pagina,
        "_indice_global_xpath": 0,
    }

    # Verificar se a tabela existe
    try:
        driver.find_element(By.XPATH, '//*[@id="resultado"]')
        print("\n  ✓ Tabela de resultados encontrada!")
    except:
        print("\n  ⚠ Tabela de resultado não encontrada para esta entidade")
        return resultado

    # 1. Extrair cabeçalhos
    resultado["headers"] = extrair_cabecalho_tabela(driver)

    pagina_atual = iniciar_de_pagina
    registro_global = 0
    indice_global_xpath = (
        0  # Mantido para checkpoint, XPath ID é calculado por página/linha
    )

    # Se iniciando de página > 1, navegar para lá
    if pagina_atual > 1:
        paginador_encontrado = False
        for xp in [
            '//*[@id="resultado_paginator_top"]',
            '//*[@id="resultado_paginator_bottom"]',
        ]:
            try:
                driver.find_element(By.XPATH, xp)
                paginador_encontrado = True
                break
            except Exception:
                continue

        if not paginador_encontrado:
            print("\n  📄 Paginador não encontrado; iniciando da página 1")
            pagina_atual = 1
        else:
            print(f"\n  ⏩ Pulando para página {pagina_atual}...")
            print(f"  📍 XPath ID começará em: {(pagina_atual - 1) * 15}")
            for _ in range(pagina_atual - 1):
                if not ir_proxima_pagina(driver):
                    print(f"  ⚠️ Não foi possível chegar até página {pagina_atual}")
                    pagina_atual = 1
                    break
                time.sleep(2)

    while True:
        print(f"\n  {'═'*60}")
        print(f"  📄 PÁGINA {pagina_atual}")
        print(f"  {'═'*60}")

        # Obter todas as linhas da página atual
        try:
            linhas = driver.find_elements(By.XPATH, '//*[@id="resultado_data"]/tr')
            total_linhas = len(linhas)
            print(f"  Encontradas {total_linhas} linhas na página {pagina_atual}")
        except:
            print("  ⚠ Nenhuma linha encontrada")
            break

        if total_linhas == 0:
            break

        for row_idx in range(total_linhas):
            if pagina_atual == iniciar_de_pagina and row_idx < iniciar_de_linha:
                continue
            xpath_id = ((pagina_atual - 1) * 15) + row_idx
            registro_global += 1
            print(f"\n  {'─'*60}")
            print(
                f"  📝 REGISTRO #{registro_global} (Pág. {pagina_atual}, Linha {row_idx + 1}/{total_linhas}) - XPath ID: {xpath_id}"
            )
            print(f"  {'─'*60}")

            # Retry logic para stale element reference
            retry_count = 0
            max_retries = 3
            dados_linha = None

            while retry_count < max_retries and dados_linha is None:
                try:
                    linha = linhas[row_idx]
                    colunas = linha.find_elements(By.TAG_NAME, "td")
                    # Pular a primeira coluna (índice 0) que é o botão de expansão
                    dados_linha = [col.text.strip() for col in colunas[1:]]
                    print(f"  Dados: {' | '.join(dados_linha[:4])}...")
                except StaleElementReferenceException:
                    retry_count += 1
                    if retry_count < max_retries:
                        print(
                            f"  ⚠ Elemento obsoleto (stale element) - Aguardando 10 segundos (tentativa {retry_count}/{max_retries})..."
                        )
                        time.sleep(10)
                        # Atualizar a lista de linhas
                        try:
                            linhas = driver.find_elements(
                                By.XPATH, '//*[@id="resultado_data"]/tr'
                            )
                            print(
                                f"  ✓ Lista de linhas atualizada. Tentando novamente..."
                            )
                        except Exception as update_error:
                            print(f"  ⚠ Erro ao atualizar linhas: {update_error}")
                    else:
                        print(
                            f"  ❌ Erro ao ler linha {row_idx} após {max_retries} tentativas"
                        )
                        continue
                except Exception as e:
                    print(f"  ⚠ Erro ao ler linha {row_idx}: {e}")
                    break

            if dados_linha is None:
                continue

            registro = {
                "dados_linha": dados_linha,
                "subtabela": {"headers": [], "linhas": []},
                "andamento": "",
                "beneficiarios": "",
            }

            # 4a. Clicar no botão de expansão (td[1])
            try:
                btn_expand = None
                if colunas:
                    try:
                        btn_expand = colunas[0].find_element(By.TAG_NAME, "div")
                    except Exception:
                        btn_expand = colunas[0]

                if btn_expand:
                    driver.execute_script(
                        "arguments[0].scrollIntoView(true);", btn_expand
                    )
                    time.sleep(0.5)
                    driver.execute_script("arguments[0].click();", btn_expand)
                    print(
                        f"    ✓ Linha expandida (índice local: {row_idx}, índice XPath: {xpath_id})"
                    )
                    time.sleep(2)

                    # 4b. Extrair subtabela (usa índice global)
                    registro["subtabela"] = extrair_subtabela(driver, xpath_id)
            except Exception as e:
                print(f"    ⚠ Erro ao expandir linha {row_idx}: {e}")

            # 4c/4d/4e. Clicar em idAndamento e extrair detalhes + beneficiários (usa índice global)
            detalhes = extrair_andamento_e_beneficiarios(driver, wait, xpath_id)
            registro["andamento"] = detalhes["andamento"]
            registro["andamento_completo"] = detalhes["andamento_completo"]
            registro["andamento_campos"] = detalhes["andamento_campos"]
            registro["beneficiarios"] = detalhes["beneficiarios"]

            resultado["registros"].append(registro)

            # Resumo do registro
            print(f"\n    📊 Resumo Registro #{registro_global}:")
            print(f"    Subtabela: {len(registro['subtabela']['linhas'])} linhas")
            print(f"    Andamento: {'✅' if registro['andamento'] else '❌'}")
            print(f"    Beneficiários: {'✅' if registro['beneficiarios'] else '❌'}")

            # Atualizar o último XPath ID calculado
            indice_global_xpath = xpath_id

        # 💾 SALVAR CHECKPOINT APÓS CADA PÁGINA
        if callback_checkpoint:
            try:
                callback_checkpoint(pagina_atual)
                print(f"\n  ✅ Checkpoint página {pagina_atual} salvo")
            except Exception as e:
                print(f"  ⚠️ Erro ao salvar checkpoint: {e}")

        # Atualizar informações de rastreamento no resultado
        resultado["_pagina_atual"] = pagina_atual
        resultado["_indice_global_xpath"] = indice_global_xpath

        # 💾 SALVAR INCREMENTAL após cada página extraída
        print(f"\n  💾 Salvando página {pagina_atual} no banco...")
        salvar_em_postgres_incremental(
            [
                {
                    "info_entidade": info_entidade,
                    "resultado": resultado,
                }
            ]
        )
        print(f"  ✅ Página {pagina_atual} salva no banco\n")

        # Tentar ir para a próxima página
        if not ir_proxima_pagina(driver):
            break
        pagina_atual += 1
        time.sleep(2)

    # Resumo final da extração
    print(f"\n{'='*70}")
    print(f"📊 RESUMO EXTRAÇÃO DA TABELA")
    print(f"{'='*70}")
    print(f"Total de páginas percorridas: {pagina_atual}")
    print(f"Total de registros extraídos: {len(resultado['registros'])}")
    print(f"Cabeçalhos: {resultado['headers']}")

    # Exibir tabela resumo dos registros
    if resultado["registros"]:
        resumo = []
        for idx, reg in enumerate(resultado["registros"], 1):
            resumo.append(
                [
                    idx,
                    " | ".join(reg["dados_linha"][:3]) if reg["dados_linha"] else "N/A",
                    len(reg["subtabela"]["linhas"]),
                    "✅" if reg["andamento"] else "❌",
                    "✅" if reg["beneficiarios"] else "❌",
                ]
            )
        print(
            tabulate(
                resumo,
                headers=[
                    "#",
                    "Dados (resumo)",
                    "Subtab. Linhas",
                    "Andamento",
                    "Beneficiários",
                ],
                tablefmt="grid",
            )
        )
    print(f"{'='*70}\n")

    return resultado


def obter_conexao_postgres():
    base_dir = Path(__file__).resolve().parent
    env_path = base_dir / ".env"
    load_dotenv(env_path)

    user = os.getenv("TJPB_DB_USER")
    password = os.getenv("TJPB_DB_PASSWORD")
    dbname = os.getenv("TJPB_DB_NAME")
    host = os.getenv("TJPB_DB_HOST")
    port = os.getenv("TJPB_DB_PORT", "5432")

    if not all([user, password, dbname, host, port]):
        raise ValueError("Credenciais incompletas no .env")

    return psycopg2.connect(
        dbname=dbname,
        user=user,
        password=password,
        host=host,
        port=port,
    )


def obter_ultimo_checkpoint_excel():
    nome_arquivo = "TJMG_Dados_INCREMENTAL.xlsx"
    if not os.path.exists(nome_arquivo):
        return None, None, None

    df = pd.read_excel(nome_arquivo)
    if df.empty:
        return None, None, None

    # Pegar último registro
    ultima_linha = df.iloc[-1]
    ent_idx = (
        int(ultima_linha.get("Entidade #", 0)) if "Entidade #" in df.columns else None
    )
    xpath_id = (
        int(ultima_linha.get("_XPath ID", 0)) if "_XPath ID" in df.columns else None
    )

    if xpath_id is not None:
        # Fórmula zero-based:
        # pagina = (xpath_id // 15) + 1
        # linha (1..15) = (xpath_id % 15) + 1
        pagina = (xpath_id // 15) + 1
        linha_na_pagina = xpath_id % 15  # 0..14

        # Se já completou 15 registros (linha 14), avança para a próxima página
        if linha_na_pagina == 14:
            return ent_idx, pagina + 1, 0

        # Caso contrário, continuar na mesma página a partir da próxima linha
        return ent_idx, pagina, linha_na_pagina + 1

    # Fallback: sem xpath_id, usar página salva e começar da primeira linha
    pagina = (
        int(ultima_linha.get("_Página Extraída", 1))
        if "_Página Extraída" in df.columns
        else 1
    )
    return ent_idx, pagina, 0


def obter_ultimo_checkpoint():
    """
    Obtém o último checkpoint de onde foi parado (entidade, página, próxima_linha).
    Retorna tupla: (ent_idx, pagina_atual, linha_inicial) ou (None, None, None) se não encontrado.
    """
    schema_name = "ordens_cronologicas"
    table_name = "tjmg"

    try:
        conn = obter_conexao_postgres()
        with conn.cursor() as cur:
            cur.execute(
                sql.SQL(
                    'SELECT "Entidade #", "_XPath ID", "_Página Extraída" '
                    "FROM {}.{} ORDER BY id DESC LIMIT 1"
                ).format(sql.Identifier(schema_name), sql.Identifier(table_name))
            )
            row = cur.fetchone()
        conn.close()

        if not row:
            return None, None, None

        ent_idx, xpath_id, pagina_extraida = row

        if xpath_id is not None:
            pagina = (xpath_id // 15) + 1
            linha_na_pagina = xpath_id % 15  # 0..14

            if linha_na_pagina == 14:
                return ent_idx, pagina + 1, 0

            return ent_idx, pagina, linha_na_pagina + 1

        if pagina_extraida is not None:
            return ent_idx, pagina_extraida, 0

        return None, None, None
    except Exception as e:
        print(f"⚠️ Erro ao ler checkpoint do banco: {e}")
        try:
            return obter_ultimo_checkpoint_excel()
        except Exception as ex:
            print(f"⚠️ Erro ao ler checkpoint do Excel: {ex}")
            return None, None, None


SQL_AUDITORIA_EXTRACAO = """
WITH base AS (
    SELECT
        t.id,
        t.criado_em,
        NULLIF(btrim(t."Ente Devedor"), '') AS ente,
        NULLIF(btrim(t."Precatório Nº"), '') AS precatorio,
        COALESCE(NULLIF(btrim(t."N° SEI"), ''), 'SEM_SEI') AS sei,
        COALESCE(NULLIF(btrim(t."Protocolo"), ''), 'SEM_PROTOCOLO') AS protocolo,
        NULLIF(btrim(t."Subtabela - 2 Ordem Aberto/Suspenso:"), '') AS ordem_txt,
        NULLIF(regexp_replace(NULLIF(btrim(t."Subtabela - 2 Ordem Aberto/Suspenso:"), ''), '\\D', '', 'g'), '')::bigint AS ordem_num,
        COALESCE(
            NULLIF(btrim(t."Subtabela - Valor de formação do Precatório (Valor de Face):"), ''),
            NULLIF(btrim(t."Valor de Face"), '')
        ) AS valor_txt,
        t."_Página Extraída" AS pagina,
        t."_XPath ID" AS xpath_id
    FROM ordens_cronologicas.tjmg t
    WHERE
        NULLIF(btrim(t."Ente Devedor"), '') IS NOT NULL
        AND NULLIF(btrim(t."Precatório Nº"), '') IS NOT NULL
        AND (%s IS NULL OR t."Entidade #" = %s)
),

rep_item AS (
    SELECT
        ente, precatorio, sei, protocolo,
        COUNT(*) AS vezes_salvo,
        COUNT(DISTINCT pagina) AS paginas_distintas,
        COUNT(DISTINCT xpath_id) AS xpaths_distintos,
        MIN(criado_em) AS criado_em_min,
        MAX(criado_em) AS criado_em_max
    FROM base
    GROUP BY 1,2,3,4
    HAVING COUNT(*) > 1
),

rep_mesmo_alvo AS (
    SELECT
        ente, precatorio, sei, protocolo,
        pagina, xpath_id,
        COUNT(*) AS vezes_no_mesmo_alvo,
        MIN(criado_em) AS criado_em_min,
        MAX(criado_em) AS criado_em_max
    FROM base
    GROUP BY 1,2,3,4,5,6
    HAVING COUNT(*) > 1
),

xpath_instavel AS (
    SELECT
        ente, precatorio, pagina, xpath_id,
        COUNT(*) AS linhas,
        COUNT(DISTINCT sei) AS seis_distintos,
        COUNT(DISTINCT protocolo) AS protocolos_distintos,
        COUNT(DISTINCT ordem_num) FILTER (WHERE ordem_num IS NOT NULL) AS ordens_distintas,
        COUNT(DISTINCT valor_txt) FILTER (WHERE valor_txt IS NOT NULL) AS valores_distintos
    FROM base
    GROUP BY 1,2,3,4
    HAVING COUNT(DISTINCT sei) > 1
            OR COUNT(DISTINCT protocolo) > 1
            OR COUNT(DISTINCT ordem_num) FILTER (WHERE ordem_num IS NOT NULL) > 1
            OR COUNT(DISTINCT valor_txt) FILTER (WHERE valor_txt IS NOT NULL) > 1
)

SELECT *
FROM (
    SELECT
        'A_RESUMO' AS secao,
        NULL::text AS ente,
        NULL::text AS precatorio,
        NULL::text AS sei,
        NULL::text AS protocolo,
        NULL::int AS pagina,
        NULL::int AS xpath_id,
        NULL::bigint AS qtd,
        jsonb_build_object(
            'total_linhas', (SELECT COUNT(*) FROM base),
            'itens_logicos_repetidos', (SELECT COUNT(*) FROM rep_item),
            'excesso_total_por_item_logico', (SELECT COALESCE(SUM(vezes_salvo - 1),0) FROM rep_item),
            'repeticoes_no_mesmo_alvo', (SELECT COUNT(*) FROM rep_mesmo_alvo),
            'excesso_total_no_mesmo_alvo', (SELECT COALESCE(SUM(vezes_no_mesmo_alvo - 1),0) FROM rep_mesmo_alvo),
            'xpaths_instaveis', (SELECT COUNT(*) FROM xpath_instavel)
        ) AS detalhes,
        1 AS ordem_secao,
        0::bigint AS ordem
    UNION ALL

    SELECT
        'B_TOP_ITENS_REPETIDOS' AS secao,
        r.ente,
        r.precatorio,
        r.sei,
        r.protocolo,
        NULL::int AS pagina,
        NULL::int AS xpath_id,
        r.vezes_salvo AS qtd,
        jsonb_build_object(
            'paginas_distintas', r.paginas_distintas,
            'xpaths_distintos', r.xpaths_distintos,
            'criado_em_min', r.criado_em_min,
            'criado_em_max', r.criado_em_max
        ) AS detalhes,
        2 AS ordem_secao,
        r.vezes_salvo AS ordem
    FROM rep_item r

    UNION ALL

    SELECT
        'C_TOP_REPETICAO_MESMO_ALVO' AS secao,
        r.ente,
        r.precatorio,
        r.sei,
        r.protocolo,
        r.pagina,
        r.xpath_id,
        r.vezes_no_mesmo_alvo AS qtd,
        jsonb_build_object(
            'criado_em_min', r.criado_em_min,
            'criado_em_max', r.criado_em_max
        ) AS detalhes,
        3 AS ordem_secao,
        r.vezes_no_mesmo_alvo AS ordem
    FROM rep_mesmo_alvo r

    UNION ALL

    SELECT
        'D_XPATH_INSTAVEL' AS secao,
        x.ente,
        x.precatorio,
        NULL::text AS sei,
        NULL::text AS protocolo,
        x.pagina,
        x.xpath_id,
        x.linhas AS qtd,
        jsonb_build_object(
            'seis_distintos', x.seis_distintos,
            'protocolos_distintos', x.protocolos_distintos,
            'ordens_distintas', x.ordens_distintas,
            'valores_distintos', x.valores_distintos
        ) AS detalhes,
        4 AS ordem_secao,
        x.linhas AS ordem
    FROM xpath_instavel x
) z
ORDER BY
    ordem_secao,
    ordem DESC NULLS LAST,
    ente NULLS LAST,
    precatorio NULLS LAST
LIMIT %s;
"""


def executar_auditoria_extracao(limite=400, entidade_num=None):
    conn = obter_conexao_postgres()
    try:
        with conn.cursor() as cur:
            cur.execute(SQL_AUDITORIA_EXTRACAO, (entidade_num, entidade_num, limite))
            return cur.fetchall()
    finally:
        conn.close()


def imprimir_resumo_auditoria(rows, entidade_num=None, ente_devedor=None):
    if not rows:
        print("  ⚠ Auditoria sem retorno de dados")
        return

    resumo = None
    top_mesmo_alvo = []
    top_xpath_instavel = []

    for row in rows:
        secao, ente, precatorio, sei, protocolo, pagina, xpath_id, qtd, detalhes = row[
            :9
        ]
        if secao == "A_RESUMO":
            resumo = detalhes or {}
        elif secao == "C_TOP_REPETICAO_MESMO_ALVO" and len(top_mesmo_alvo) < 3:
            top_mesmo_alvo.append(
                (ente, precatorio, sei, protocolo, pagina, xpath_id, qtd)
            )
        elif secao == "D_XPATH_INSTAVEL" and len(top_xpath_instavel) < 3:
            top_xpath_instavel.append(
                (ente, precatorio, pagina, xpath_id, qtd, detalhes or {})
            )

    print(f"\n{'='*70}")
    print("🧪 AUDITORIA DE EXTRAÇÃO")
    if entidade_num is not None:
        print(f"Entidade processada: #{entidade_num}")
    if ente_devedor:
        print(f"Ente devedor: {ente_devedor}")

    if not resumo:
        print("Resumo indisponível")
        print(f"{'='*70}\n")
        return

    repeticoes_mesmo_alvo = resumo.get("repeticoes_no_mesmo_alvo", 0)
    status = "❌ ERRO CONFIRMADO" if repeticoes_mesmo_alvo > 0 else "✅ SEM ERRO FORTE"
    print(f"Status: {status}")
    print(
        " | ".join(
            [
                f"total_linhas={resumo.get('total_linhas', 0)}",
                f"itens_logicos_repetidos={resumo.get('itens_logicos_repetidos', 0)}",
                f"repeticoes_no_mesmo_alvo={repeticoes_mesmo_alvo}",
                f"xpaths_instaveis={resumo.get('xpaths_instaveis', 0)}",
            ]
        )
    )

    if top_mesmo_alvo:
        print("\nTop repetições no mesmo alvo:")
        for idx, item in enumerate(top_mesmo_alvo, 1):
            ente, precatorio, sei, protocolo, pagina, xpath_id, qtd = item
            print(
                f"  {idx}) qtd={qtd} | pág={pagina} | xpath={xpath_id} | precatório={precatorio} | sei={sei} | protocolo={protocolo}"
            )

    if top_xpath_instavel:
        print("\nTop xpaths instáveis:")
        for idx, item in enumerate(top_xpath_instavel, 1):
            ente, precatorio, pagina, xpath_id, qtd, detalhes = item
            print(
                f"  {idx}) linhas={qtd} | pág={pagina} | xpath={xpath_id} | precatório={precatorio} | seis={detalhes.get('seis_distintos', 0)} | protocolos={detalhes.get('protocolos_distintos', 0)}"
            )

    print(f"{'='*70}\n")


def salvar_em_postgres_incremental(dados_todas_entidades):
    """
    Salva os dados no PostgreSQL (append incremental).
    """
    schema_name = "ordens_cronologicas"
    table_name = "tjmg"
    colunas = [
        "Entidade #",
        "Ente Devedor Principal",
        "Ente Devedor",
        "Regime de Pagamento",
        "Lei Pequeno Valor",
        "Ordem Crono. Disponível",
        "Registro #",
        "_Página Extraída",
        "_XPath ID",
        "Precatório Nº",
        "Venc.",
        "Natureza",
        "N° SEI",
        "Nº Eproc 2G",
        "Credor (Principal)",
        "Protocolo",
        "Situação",
        "Nº Processo Execução",
        "Tem Subtabela",
        "Subtabela - 1 Ordem Cronológica:",
        "Subtabela - 2 Ordem Aberto/Suspenso:",
        "Subtabela - Valor de formação do Precatório (Valor de Face):",
        "Subtabela - Processo de Execução nº:",
        "Subtabela - Processo SEI nº:",
        "Subtabela - Processo Eproc 2ª Instância nº:",
        "Subtabela - Origem:",
        "Subtabela - Ação:",
        "Subtabela - Liquidação:",
        "Subtabela - Protocolo (Data/Hora):",
        "Subtabela - Protocolo (Número/Ano):",
        "Credor Principal",
        "Número e Natureza do Precatório",
        "Ano de Vencimento",
        "Processo Eproc 2ª Instância",
        "Valor de Face",
        "Data de Liquidação",
        "Protocolo Data/Hora",
        "Protocolo Número/Ano",
        "Processo de Execução",
        "Processo SEI",
        "Origem",
        "Ação",
        "Beneficiários",
        "Subtabela - Linhas",
    ]

    registros_para_inserir = []

    for ent_idx, entidade_data in enumerate(dados_todas_entidades, 1):
        info = entidade_data.get("info_entidade", {})
        resultado = entidade_data.get("resultado", {})
        registros = resultado.get("registros", [])

        if not registros:
            continue

        headers_tabela = resultado.get("headers", [])

        for reg_idx, reg in enumerate(registros, 1):
            dados_linha = reg.get("dados_linha", [])

            linha_consolidada = {
                "Entidade #": ent_idx,
                "Ente Devedor Principal": info.get("Ente Devedor Principal", "N/A"),
                "Ente Devedor": info.get("Ente Devedor", "N/A"),
                "Regime de Pagamento": info.get("Regime de Pagamento", "N/A"),
                "Lei Pequeno Valor": info.get("Lei Pequeno Valor", "N/A"),
                "Ordem Crono. Disponível": not info.get(
                    "ordem_cronologica_nao_disponivel", False
                ),
                "Registro #": reg_idx,
                "_Página Extraída": resultado.get("_pagina_atual", 1),
                "_XPath ID": resultado.get("_indice_global_xpath", 0),
            }

            for col_idx, header in enumerate(headers_tabela):
                if col_idx < len(dados_linha):
                    linha_consolidada[header] = dados_linha[col_idx]

            linha_consolidada["Tem Subtabela"] = bool(
                reg.get("subtabela", {}).get("linhas")
            )

            sub = reg.get("subtabela", {})
            sub_headers = sub.get("headers", [])
            sub_linhas = sub.get("linhas", [])

            dados_kv_sub = parsear_subtabela_linhas(sub_linhas)

            if dados_kv_sub:
                for chave, valor in dados_kv_sub.items():
                    linha_consolidada[f"Subtabela - {chave}"] = valor
            elif sub_headers:
                for idx_header, header in enumerate(sub_headers):
                    valores_coluna = []
                    for linha_sub in sub_linhas:
                        if idx_header < len(linha_sub):
                            valor = linha_sub[idx_header].strip()
                            if valor:
                                valores_coluna.append(valor)

                    linha_consolidada[f"Subtabela - {header}"] = " | ".join(
                        valores_coluna
                    )
            else:
                if sub_linhas:
                    linhas_compactadas = ["; ".join(l) for l in sub_linhas]
                    linha_consolidada["Subtabela - Linhas"] = " | ".join(
                        linhas_compactadas
                    )
                else:
                    linha_consolidada["Subtabela - Linhas"] = ""

            campos_andamento_padrao = [
                "Credor Principal",
                "Número e Natureza do Precatório",
                "Ano de Vencimento",
                "Processo Eproc 2ª Instância",
                "Situação",
                "Valor de Face",
                "Data de Liquidação",
                "Protocolo Data/Hora",
                "Protocolo Número/Ano",
                "Processo de Execução",
                "Processo SEI",
                "Origem",
                "Ação",
            ]

            andamento_campos = reg.get("andamento_campos", {})
            for campo_nome in campos_andamento_padrao:
                linha_consolidada[campo_nome] = andamento_campos.get(campo_nome, "")

            beneficiarios = reg.get("beneficiarios", "")
            if isinstance(beneficiarios, list):
                beneficiarios = " | ".join([b for b in beneficiarios if b])
            linha_consolidada["Beneficiários"] = beneficiarios or ""

            if "Subtabela - Linhas" not in linha_consolidada:
                linha_consolidada["Subtabela - Linhas"] = ""

            registros_para_inserir.append(
                [linha_consolidada.get(coluna) for coluna in colunas]
            )

    if not registros_para_inserir:
        return 0

    conn = obter_conexao_postgres()
    conn.autocommit = True
    insert_sql = sql.SQL("INSERT INTO {}.{} ({}) VALUES ({})").format(
        sql.Identifier(schema_name),
        sql.Identifier(table_name),
        sql.SQL(", ").join([sql.Identifier(c) for c in colunas]),
        sql.SQL(", ").join(sql.Placeholder() for _ in colunas),
    )

    with conn.cursor() as cur:
        execute_batch(cur, insert_sql, registros_para_inserir, page_size=200)

    conn.close()
    return len(registros_para_inserir)


def salvar_em_excel_incremental(dados_todas_entidades):
    """
    Salva os dados de forma incremental (acumula ao arquivo existente).
    """
    import os

    nome_arquivo = "TJMG_Dados_INCREMENTAL.xlsx"
    caminho_completo = os.path.abspath(nome_arquivo)
    print(f"📁 Salvando em: {caminho_completo}")
    return salvar_em_excel_com_append(dados_todas_entidades, nome_arquivo)


def salvar_em_excel_com_append(dados_todas_entidades, nome_arquivo):
    """
    Salva os dados em Excel fazendo append se o arquivo já existir.
    """
    import os

    caminho_completo = os.path.abspath(nome_arquivo)

    # Ler dados anteriores se o arquivo existir
    dados_anteriores = []
    if os.path.exists(nome_arquivo):
        try:
            df_anterior = pd.read_excel(nome_arquivo)
            dados_anteriores = df_anterior.to_dict("records")
            print(
                f"  📖 Arquivo anterior encontrado com {len(dados_anteriores)} registros"
            )
        except Exception as e:
            print(f"  ⚠️ Não foi possível ler arquivo anterior: {e}")

    print(f"\n{'='*70}")
    print(f"💾 SALVANDO DADOS EM EXCEL: {nome_arquivo}")
    print(f"📁 Caminho completo: {caminho_completo}")
    print(f"📊 Total de entidades a salvar: {len(dados_todas_entidades)}")
    print(f"{'='*70}")

    try:
        # Consolidar TODOS os dados em uma ÚNICA planilha
        dados_consolidados = dados_anteriores.copy()

        for ent_idx, entidade_data in enumerate(dados_todas_entidades, 1):
            info = entidade_data.get("info_entidade", {})
            resultado = entidade_data.get("resultado", {})
            registros = resultado.get("registros", [])

            if not registros:
                continue

            headers_tabela = resultado.get("headers", [])

            for reg_idx, reg in enumerate(registros, 1):
                dados_linha = reg.get("dados_linha", [])

                # Criar linha consolidada com TODOS os campos
                linha_consolidada = {
                    "Entidade #": ent_idx,
                    "Ente Devedor Principal": info.get("Ente Devedor Principal", "N/A"),
                    "Ente Devedor": info.get("Ente Devedor", "N/A"),
                    "Regime de Pagamento": info.get("Regime de Pagamento", "N/A"),
                    "Lei Pequeno Valor": info.get("Lei Pequeno Valor", "N/A"),
                    "Ordem Crono. Disponível": (
                        "Não" if info.get("ordem_cronologica_nao_disponivel") else "Sim"
                    ),
                    "Registro #": reg_idx,
                    "_Página Extraída": resultado.get("_pagina_atual", 1),
                    "_XPath ID": resultado.get("_indice_global_xpath", 0),
                }

                # Adicionar todas as colunas da tabela
                for col_idx, header in enumerate(headers_tabela):
                    if col_idx < len(dados_linha):
                        linha_consolidada[header] = dados_linha[col_idx]
                    else:
                        linha_consolidada[header] = ""

                # Adicionar campos adicionais
                linha_consolidada["Tem Subtabela"] = (
                    "Sim" if reg.get("subtabela", {}).get("linhas") else "Não"
                )

                # Adicionar colunas da subtabela (uma coluna por header)
                sub = reg.get("subtabela", {})
                sub_headers = sub.get("headers", [])
                sub_linhas = sub.get("linhas", [])

                dados_kv_sub = parsear_subtabela_linhas(sub_linhas)

                if dados_kv_sub:
                    for chave, valor in dados_kv_sub.items():
                        linha_consolidada[f"Subtabela - {chave}"] = valor
                elif sub_headers:
                    for idx_header, header in enumerate(sub_headers):
                        valores_coluna = []
                        for linha_sub in sub_linhas:
                            if idx_header < len(linha_sub):
                                valor = linha_sub[idx_header].strip()
                                if valor:
                                    valores_coluna.append(valor)

                        linha_consolidada[f"Subtabela - {header}"] = " | ".join(
                            valores_coluna
                        )
                else:
                    if sub_linhas:
                        linhas_compactadas = ["; ".join(l) for l in sub_linhas]
                        linha_consolidada["Subtabela - Linhas"] = " | ".join(
                            linhas_compactadas
                        )
                    else:
                        linha_consolidada["Subtabela - Linhas"] = ""

                # Adicionar campos parseados do andamento (cada campo em uma coluna)
                # Garantir que todas as colunas existam mesmo se vazias
                campos_andamento_padrao = [
                    "Credor Principal",
                    "Número e Natureza do Precatório",
                    "Ano de Vencimento",
                    "Processo Eproc 2ª Instância",
                    "Situação",
                    "Valor de Face",
                    "Data de Liquidação",
                    "Protocolo Data/Hora",
                    "Protocolo Número/Ano",
                    "Processo de Execução",
                    "Processo SEI",
                    "Origem",
                    "Ação",
                ]

                andamento_campos = reg.get("andamento_campos", {})
                for campo_nome in campos_andamento_padrao:
                    linha_consolidada[campo_nome] = andamento_campos.get(campo_nome, "")

                # Adicionar beneficiários
                benef_text = reg.get("beneficiarios", "")
                linha_consolidada["Beneficiários"] = benef_text if benef_text else ""

                dados_consolidados.append(linha_consolidada)

        # Salvar em arquivo Excel único com uma página
        if dados_consolidados:
            df_consolidado = pd.DataFrame(dados_consolidados)
            with pd.ExcelWriter(nome_arquivo, engine="openpyxl") as writer:
                df_consolidado.to_excel(
                    writer, sheet_name="Dados Completos", index=False
                )
                print(
                    f"  ✓ Aba 'Dados Completos' criada com {len(dados_consolidados)} registros"
                )

        # Aplicar formatação
        aplicar_formatacao_excel(nome_arquivo)

        print(f"\n✅ Arquivo Excel salvo com sucesso!")
        print(f"📁 Localização: {caminho_completo}")
        print(f"{'='*70}\n")

        return nome_arquivo

    except Exception as e:
        import traceback

        print(f"\n❌ Erro ao salvar Excel: {e}")
        print(f"Stack trace completo:")
        traceback.print_exc()
        return None


def aplicar_formatacao_excel(nome_arquivo):
    """Aplica formatação visual ao arquivo Excel."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(nome_arquivo)

        # Estilo para cabeçalhos
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Formatar cabeçalhos (primeira linha)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

            # Ajustar largura das colunas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Congelar primeira linha
            ws.freeze_panes = ws["A2"]

        wb.save(nome_arquivo)
        print("  ✓ Formatação aplicada com sucesso")

    except Exception as e:
        print(f"  ⚠ Não foi possível aplicar formatação: {e}")


def main():
    # Configurar o WebDriver do Chrome
    print("Iniciando o navegador...")

    try:
        # Adicionar opções do Chrome
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")

        # Tentar inicializar o Chrome sem especificar o driver
        # (o Selenium 4.6+ consegue gerenciar automaticamente)
        driver = webdriver.Chrome(options=chrome_options)
    except Exception as e:
        print(f"Erro ao inicializar com gerenciamento automático: {e}")
        print("Tentando com webdriver-manager...")

        # Forçar novo download com webdriver-manager
        from webdriver_manager.chrome import ChromeDriverManager

        driver_path = ChromeDriverManager(cache_valid_range=0).install()
        print(f"Driver baixado em: {driver_path}")

        driver = webdriver.Chrome(service=Service(driver_path), options=chrome_options)

    try:
        # Maximizar a janela
        driver.maximize_window()

        # Acessar o site do TJMG
        url = "https://www8.tjmg.jus.br/juridico/pe/listaCronologia.jsf"
        print(f"Acessando: {url}")
        driver.get(url)

        # Aguardar o carregamento da página
        wait = WebDriverWait(driver, 10)
        print("Página carregada com sucesso!")

        # Aguardar a página carregar completamente
        time.sleep(3)

        # Clicar no botão dropdown para abrir o modal
        print("Procurando o botão dropdown...")
        dropdown_button = wait.until(
            EC.element_to_be_clickable(
                (
                    By.CSS_SELECTOR,
                    "span.ui-button-icon-primary.ui-icon.ui-icon-triangle-1-s",
                )
            )
        )
        print("Clicando no dropdown...")
        dropdown_button.click()

        # Aguardar o modal abrir e estar visível
        time.sleep(2)

        # Encontrar todas as entidades na lista usando o XPath correto
        print("Buscando lista de entidades...")
        wait.until(
            EC.presence_of_element_located(
                (By.XPATH, '//*[@id="entidade_devedora_panel"]/ul')
            )
        )

        # Contar quantas entidades existem
        entidades = driver.find_elements(
            By.XPATH, '//*[@id="entidade_devedora_panel"]/ul/li'
        )
        total_entidades = len(entidades)
        print(f"Encontradas {total_entidades} entidades")

        # ✅ Verificar último checkpoint
        ent_checkpoint, pagina_checkpoint, linha_inicio_checkpoint = (
            obter_ultimo_checkpoint()
        )
        if ent_checkpoint is not None:
            print(f"\n{'='*70}")
            print(f"✅ CHECKPOINT ENCONTRADO")
            print(f"Última entidade processada: #{ent_checkpoint}")
            print(f"Última página processada: {pagina_checkpoint}")
            print(
                f"Será retomada a partir da entidade #{ent_checkpoint} página {pagina_checkpoint}"
            )
            print(f"{'='*70}\n")
        else:
            print(f"\n📍 Iniciando nova varredura (sem checkpoint anterior)\n")
            linha_inicio_checkpoint = 0

        # Lista para armazenar dados de todas as entidades
        dados_todas_entidades = []

        # Clicar em cada entidade uma por uma
        for i in range(total_entidades):
            try:
                # Reabrir o dropdown antes de cada clique
                if i > 0:
                    print("\nReabrindo o dropdown...")
                    dropdown_button = wait.until(
                        EC.element_to_be_clickable(
                            (
                                By.CSS_SELECTOR,
                                "span.ui-button-icon-primary.ui-icon.ui-icon-triangle-1-s",
                            )
                        )
                    )
                    dropdown_button.click()
                    time.sleep(1.5)

                # Aguardar a lista estar visível
                wait.until(
                    EC.visibility_of_element_located(
                        (By.XPATH, '//*[@id="entidade_devedora_panel"]/ul')
                    )
                )

                # Buscar novamente os elementos (para evitar stale reference)
                entidades = driver.find_elements(
                    By.XPATH, '//*[@id="entidade_devedora_panel"]/ul/li'
                )

                if i < len(entidades):
                    entidade = entidades[i]
                    entidade_texto = entidade.text

                    # Scroll até o elemento se necessário
                    driver.execute_script(
                        "arguments[0].scrollIntoView(true);", entidade
                    )
                    time.sleep(0.5)

                    print(
                        f"\n[{i+1}/{total_entidades}] Clicando na entidade: {entidade_texto}"
                    )

                    # Tentar clicar usando JavaScript se o click normal falhar
                    try:
                        wait.until(EC.element_to_be_clickable(entidade))
                        entidade.click()
                    except:
                        driver.execute_script("arguments[0].click();", entidade)

                    # Aguardar o dropdown fechar
                    time.sleep(1.5)

                    # Tentar apertar Enter ou clicar no botão Consultar
                    print("Confirmando com Enter/Consultar...")
                    try:
                        # Primeiro tenta pressionar Enter
                        ActionChains(driver).send_keys(Keys.ENTER).perform()
                        print("Enter pressionado")
                    except:
                        try:
                            # Se falhar, tenta clicar no botão Consultar
                            botao_consulta = wait.until(
                                EC.element_to_be_clickable(
                                    (By.XPATH, '//*[@id="consulta2"]')
                                )
                            )
                            botao_consulta.click()
                            print("Botão Consultar clicado")
                        except:
                            # Última tentativa: usar JavaScript
                            botao_consulta = driver.find_element(
                                By.XPATH, '//*[@id="consulta2"]'
                            )
                            driver.execute_script(
                                "arguments[0].click();", botao_consulta
                            )
                            print("Botão Consultar clicado via JavaScript")

                    # Aguardar a página processar a consulta
                    print("Aguardando processamento...")
                    time.sleep(4)

                    # ✨ EXTRAIR INFORMAÇÕES DO MODAL ABERTO
                    print(f"\n🔍 Extraindo informações da entidade {i+1}...")
                    informacoes = extrair_informacoes_modal(driver)
                    exibir_informacoes_tabular(informacoes, i + 1)

                    # ✨ EXTRAIR TOTAL DE PRECATÓRIOS PARA VALIDAÇÃO
                    print(f"\n🔍 Extraindo total de precatórios para validação...")
                    total_precatorios = extrair_total_precatorios(driver)
                    if total_precatorios is not None:
                        salvar_validacao(
                            entidade_num=i + 1,
                            ente_devedor_principal=informacoes.get(
                                "Ente Devedor Principal", ""
                            ),
                            ente_devedor=informacoes.get("Ente Devedor", ""),
                            total_precatorios=total_precatorios,
                            worker_id=None,
                        )

                    # ✨ EXTRAIR DADOS DA TABELA DE RESULTADO
                    dados_tabela_resultado = {}
                    pagina_inicio = 1
                    linha_inicio = 0
                    if not informacoes.get("ordem_cronologica_nao_disponivel"):
                        print(
                            f"\n📋 Extraindo tabela de resultados da entidade {i+1}..."
                        )
                        # Se for a entidade do checkpoint, começar da página onde parou
                        if ent_checkpoint is not None and (i + 1) == ent_checkpoint:
                            pagina_inicio = pagina_checkpoint
                            linha_inicio = linha_inicio_checkpoint
                            print(
                                f"  ⏩ Retomando a partir da página {pagina_inicio} (linha {linha_inicio + 1})"
                            )
                        dados_tabela_resultado = extrair_tabela_resultado(
                            driver,
                            wait,
                            informacoes,
                            iniciar_de_pagina=pagina_inicio,
                            iniciar_de_linha=linha_inicio,
                        )
                    else:
                        print(
                            f"\n⏩ Pulando extração da tabela "
                            f"(ordem cronológica não disponível)"
                        )
                        dados_tabela_resultado = {"headers": [], "registros": []}

                    # Armazenar dados desta entidade
                    dados_todas_entidades.append(
                        {
                            "info_entidade": informacoes,
                            "resultado": dados_tabela_resultado,
                        }
                    )

                    try:
                        rows_auditoria = executar_auditoria_extracao(limite=400)
                        imprimir_resumo_auditoria(
                            rows_auditoria,
                            entidade_num=i + 1,
                            ente_devedor=informacoes.get("Ente Devedor", ""),
                        )
                    except Exception as auditoria_error:
                        print(
                            f"  ⚠ Falha ao executar auditoria de extração: {auditoria_error}"
                        )

                    # Obs: Salvamento incremental já acontece após cada linha extraída
                    # Não é necessário salvar novamente por entidade

                    # Voltar para o site principal antes da próxima entidade
                    if i < total_entidades - 1:
                        print("Voltando para o site principal...")
                        driver.get(url)
                        time.sleep(3)

            except Exception as e:
                print(f"Erro ao clicar na entidade {i+1}: {e}")
                # Voltar para o site principal em caso de erro
                try:
                    driver.get(url)
                    time.sleep(3)
                except:
                    pass
                continue

        print("\nTodas as entidades foram testadas!")

        # 💾 SALVAR TODOS OS DADOS NO BANCO
        if dados_todas_entidades:
            print(f"\n{'='*70}")
            print(f"📊 Total de entidades processadas: {len(dados_todas_entidades)}")
            print(f"{'='*70}")

            total_inserido = salvar_em_postgres_incremental(dados_todas_entidades)

            if total_inserido:
                print(f"\n🎉 PROCESSO CONCLUÍDO COM SUCESSO!")
                print(f"🗄️ Registros inseridos: {total_inserido}")
        else:
            print("\n⚠ Nenhum dado foi coletado para salvar.")

        time.sleep(3)

    except Exception as e:
        print(f"Erro durante a automação: {e}")

    finally:
        # Fechar o navegador
        print("Fechando o navegador...")
        driver.quit()


if __name__ == "__main__":
    main()
